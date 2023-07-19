import boto3
from openpyxl import Workbook

#Creating an Excel workbook
wb = Workbook()
wb.create_sheet("Sheet_one")

#Creating Column Names in One Sheet
ws1 = wb['Sheet_one']
ws1['A2'] = 'LoadBalncer Name'
ws1['B2'] = 'Bucket Path'

Profile=input('Enter your Profile name')
Region=input('Enter the Region where you want to get the Loadbalancer data from:')

boto3.setup_default_session(profile_name=Profile, region_name=Region)

def get_all_load_balancer_access_logs_buckets():
    try:
        # Initialize the AWS SDK client for Elastic Load Balancing
        elbv2_client = boto3.client('elbv2')

        load_balancer_info_list = []

        # Describe all the load balancers in the account
        response = elbv2_client.describe_load_balancers()

        for lb_info in response['LoadBalancers']:
            lb_name = lb_info['LoadBalancerName']
            load_balancer_info_list.append(lb_name)

        return load_balancer_info_list

    except Exception as e:
        print("Error: ", e)
        return None

# Call the function to get all the load balancers and their access log S3 bucket paths
load_balancer_namest = get_all_load_balancer_access_logs_buckets()

def get_load_balancer_access_logs_buckets(load_balancer_names):
    try:
        # Initialize the AWS SDK client for Elastic Load Balancing
        elbv2_client = boto3.client('elbv2')

        access_logs_buckets = {}

        # Describe the load balancers to get their attributes
        response = elbv2_client.describe_load_balancers(Names=load_balancer_names)

        for lb_info in response['LoadBalancers']:
            lb_name = lb_info['LoadBalancerName']
            lb_arn=lb_info['LoadBalancerArn']
            response2=elbv2_client.describe_load_balancer_attributes(LoadBalancerArn=lb_arn)
            access_logs_s3_bucket = response2['Attributes']
            access_logs_buckets[lb_name] = access_logs_s3_bucket

        return access_logs_buckets

    except Exception as e:
        print("Error: ", e)
        return None
# I am using 2 for loops here. Optimization is welcome here.
for i in range(0,len(load_balancer_namest)):
    A=[]
    A.append(load_balancer_namest[i])
    access_logs_buckets = get_load_balancer_access_logs_buckets(A)

    if access_logs_buckets:
        for lb_name, s3_bucket_path in access_logs_buckets.items():
            path_dict=s3_bucket_path[2]
            bucket_dict=s3_bucket_path[1]
            row_start = 3  #start below the header row 2
            col_start = 2  #starts from column B
            print(f"Load Balancer:'{lb_name}'")
            print(f"Access logs S3 bucket path:{bucket_dict['Value']}/{path_dict['Value']}/")
            ws1.cell(row_start+i, col_start-1).value = lb_name
            ws1.cell(row_start+i, col_start).value = f"{bucket_dict['Value']}/{path_dict['Value']}/"
    else:
        print("Failed to fetch the access logs S3 bucket paths.")
#Saving the Excel Sheet after all the cells are updated.
wb.save('LoadBalancerPaths.xlsx')
